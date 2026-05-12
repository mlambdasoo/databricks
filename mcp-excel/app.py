"""
sudong-excel-mcp — FastMCP wrapper for Excel structure inspection on Databricks Apps.

UC Volume FUSE mount은 Databricks Apps 컨테이너에서 지원되지 않으므로,
이 wrapper는 Databricks SDK 를 사용해 Volume 파일을 로컬 /tmp 에 lazy 다운로드한 후
openpyxl 기반의 excel-mcp-server 내부 비즈니스 로직을 재사용한다.

Tools:
  - list_sheet_names            시트 이름 목록
  - get_workbook_metadata       workbook 전체 메타데이터 (시트, 크기, used range)
  - get_merged_cells            시트의 병합 셀 목록
  - read_data_from_excel        범위 데이터 읽기 (preview 모드 지원)
  - get_data_validation_info    데이터 검증 규칙
  - validate_formula_syntax     Excel 수식 문법 검증 (적용 X)
  - validate_excel_range        지정한 범위가 시트 내 유효한지 검증
  - detect_headers_and_types    헤더/타입 자동 추정 (openpyxl 직접 구현)
  - profile_sheet               시트 심층 프로파일링
  - suggest_delta_schema        Delta CREATE TABLE 제안

입력 filepath 는 반드시 UC Volume 절대경로 (/Volumes/...) 이어야 한다.
"""

from __future__ import annotations

import asyncio
import contextvars
import hashlib
import logging
import os
import re
import sys
import tempfile
from pathlib import Path
from typing import Any

from fastmcp import FastMCP

# excel-mcp-server 내부 모듈 재사용 (openpyxl 기반, FastMCP tool 래핑과 독립적).
# 원본 read tool 6개 (haris-musa server.py)를 1:1 로 노출하기 위한 import 묶음.
from excel_mcp.data import read_excel_range
from excel_mcp.sheet import get_merged_ranges as _get_merged_ranges_impl
from excel_mcp.cell_validation import get_all_validation_ranges as _get_all_validation_ranges_impl
from excel_mcp.validation import (
    validate_formula_in_cell_operation as _validate_formula_impl,
    validate_range_in_sheet_operation as _validate_range_impl,
)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Databricks SDK
from databricks.sdk import WorkspaceClient

# Starlette middleware:
#   CORSMiddleware   — 브라우저 기반 MCP 클라이언트(Genie Code 등)의 preflight 처리
#   UserTokenMiddleware — Databricks Apps 가 주입하는 X-Forwarded-Access-Token 추출
from starlette.middleware import Middleware
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=os.environ.get("LOG_LEVEL", "INFO"),
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
    stream=sys.stderr,
)
logger = logging.getLogger("sudong-excel-mcp")


# ---------------------------------------------------------------------------
# Globals
# ---------------------------------------------------------------------------
CACHE_DIR = Path(os.environ.get("EXCEL_CACHE_DIR", "/tmp/excel_cache"))
CACHE_DIR.mkdir(parents=True, exist_ok=True)

VOLUME_PREFIX = "/Volumes/"

# mtime 비교 tolerance (초). SDK 가 반환하는 last_modified 와 로컬 mtime 가
# 프로토콜 왕복 중 미세하게 차이 날 수 있어 1초 여유를 준다.
MTIME_TOLERANCE_SECONDS = 1.0


# ---------------------------------------------------------------------------
# OBO (On-Behalf-Of) — 요청마다 사용자 토큰 기반 WorkspaceClient 생성
# ---------------------------------------------------------------------------
# Databricks Apps 는 user_api_scopes 가 선언된 앱에 한해 요청 헤더에
# X-Forwarded-Access-Token 으로 사용자 OAuth 토큰을 주입한다.
# 이 토큰으로 Files API 등을 호출하면 앱 SP 가 아닌 "end user 본인" 권한이
# 적용되므로, Volume 에 대해 사용자 본인의 UC grants 가 그대로 반영된다.

_user_token_ctx: contextvars.ContextVar[str | None] = contextvars.ContextVar(
    "mcp_excel_user_token", default=None
)


class UserTokenMiddleware(BaseHTTPMiddleware):
    """X-Forwarded-Access-Token 헤더를 꺼내 ContextVar 에 저장한다."""

    async def dispatch(self, request, call_next):
        token = request.headers.get("x-forwarded-access-token")
        if token:
            _user_token_ctx.set(token)
        else:
            _user_token_ctx.set(None)
        return await call_next(request)


def get_user_ws() -> WorkspaceClient:
    """현재 요청의 사용자 토큰으로 인증된 WorkspaceClient 를 반환한다.

    user_api_scopes 가 선언된 앱이므로 X-Forwarded-Access-Token 이 반드시 주입된다.
    auth_type="pat" 명시: 컨테이너에 자동 주입되는 SP credential 환경변수
    (DATABRICKS_CLIENT_ID/SECRET) 를 SDK 가 함께 인식해 충돌하는 것을 방지.
    """
    token = _user_token_ctx.get()
    host = os.environ.get("DATABRICKS_HOST")
    if not (token and host):
        raise RuntimeError(
            "user OAuth token is required. "
            "X-Forwarded-Access-Token header missing — check user_api_scopes in app.yaml."
        )
    return WorkspaceClient(host=host, token=token, auth_type="pat")


# ---------------------------------------------------------------------------
# Path resolver — UC Volume 경로 → 로컬 캐시 경로 (lazy download)
# ---------------------------------------------------------------------------
def _cache_path_for(volume_path: str) -> Path:
    """Volume 경로에 대한 결정론적 로컬 캐시 경로."""
    h = hashlib.sha256(volume_path.encode("utf-8")).hexdigest()[:16]
    basename = os.path.basename(volume_path) or "file.xlsx"
    safe_basename = re.sub(r"[^A-Za-z0-9._-]+", "_", basename)
    return CACHE_DIR / f"{h}_{safe_basename}"


def _remote_mtime(volume_path: str) -> float | None:
    """Volume 파일의 서버측 mtime(초). 실패 시 None."""
    try:
        meta = get_user_ws().files.get_metadata(volume_path)
    except Exception as exc:  # 존재하지 않거나 권한 없을 때
        logger.warning("get_metadata failed for %s: %s", volume_path, exc)
        return None
    lm = getattr(meta, "last_modified", None)
    if lm is None:
        return None
    # last_modified 는 datetime 또는 epoch float 일 수 있음 — 둘 다 대응
    if hasattr(lm, "timestamp"):
        try:
            return float(lm.timestamp())
        except Exception:
            return None
    try:
        return float(lm)
    except Exception:
        return None


def _download_to(volume_path: str, local_path: Path) -> None:
    """Volume 파일을 로컬로 내려받는다 (atomic write)."""
    tmp = local_path.with_suffix(local_path.suffix + ".part")
    try:
        resp = get_user_ws().files.download(volume_path)
        contents = resp.contents  # BinaryIO (StreamingResponse)
        with open(tmp, "wb") as f:
            if hasattr(contents, "read"):
                while True:
                    chunk = contents.read(1024 * 1024)
                    if not chunk:
                        break
                    f.write(chunk)
            else:
                f.write(contents)
        os.replace(tmp, local_path)
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except Exception:
                pass


def resolve(filepath: str) -> str:
    """UC Volume 절대 경로를 받아 로컬 캐시 경로를 반환.

    - 로컬 캐시가 없거나 stale(mtime 비교)이면 재다운로드
    - /Volumes/ 로 시작하지 않는 경로는 거부
    """
    if not filepath or not filepath.startswith(VOLUME_PREFIX):
        raise ValueError(
            f"filepath must be a UC Volume absolute path starting with '{VOLUME_PREFIX}'. "
            f"got: {filepath!r}"
        )
    local = _cache_path_for(filepath)
    remote_mtime = _remote_mtime(filepath)

    need_download = False
    if not local.exists():
        need_download = True
        reason = "missing local cache"
    elif remote_mtime is None:
        # 메타데이터 확인 실패 — 기존 캐시 사용 (offline/permission 이슈일 가능성)
        reason = "metadata unavailable, using cached copy"
    else:
        local_mtime = local.stat().st_mtime
        if remote_mtime - local_mtime > MTIME_TOLERANCE_SECONDS:
            need_download = True
            reason = f"stale cache (remote={remote_mtime} > local={local_mtime})"
        else:
            reason = "cache hit"

    if need_download:
        logger.info("resolve: %s — %s → downloading", filepath, reason)
        _download_to(filepath, local)
        if remote_mtime:
            try:
                os.utime(local, (remote_mtime, remote_mtime))
            except Exception:
                pass
    else:
        logger.debug("resolve: %s — %s", filepath, reason)

    return str(local)


# ---------------------------------------------------------------------------
# MCP Server
# ---------------------------------------------------------------------------
mcp = FastMCP(os.environ.get("MCP_SERVER_NAME", "excel-mcp"))


_VOLUME_NOTE = (
    "NOTE: filepath 는 반드시 UC Volume 절대 경로여야 함 "
    "(예: /Volumes/<catalog>/<schema>/<volume>/<file>.xlsx). "
    "로컬 경로나 workspace 경로는 지원하지 않음."
)


def _get_workbook_metadata_impl(filepath: str) -> dict[str, Any]:
    """Workbook 메타 정보 추출 (내부 헬퍼). inspect_workbook 에서 사용."""
    local = resolve(filepath)

    def _iso(dt):
        try:
            return dt.isoformat() if dt else None
        except Exception:
            return None

    file_stat = os.stat(local)
    wb = load_workbook(local, read_only=False, data_only=False)
    try:
        props = wb.properties

        sheets_info: list[dict[str, Any]] = []
        for name in wb.sheetnames:
            ws = wb[name]
            merged = [str(r) for r in ws.merged_cells.ranges]
            dv = getattr(ws, "data_validations", None)
            dv_count = len(dv.dataValidation) if dv is not None else 0
            tables = getattr(ws, "tables", None)
            # openpyxl 3.1+ 에서 ws.tables 는 dict-like
            try:
                table_count = len(tables) if tables is not None else 0
            except TypeError:
                table_count = 0
            sheets_info.append({
                "name": name,
                "used_range": ws.dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "hidden": ws.sheet_state != "visible",
                "sheet_state": ws.sheet_state,
                "merged_cells": merged,
                "merged_cell_count": len(merged),
                "has_charts": bool(getattr(ws, "_charts", None)),
                "has_pivot_tables": bool(getattr(ws, "_pivots", None)),
                "table_count": table_count,
                "data_validation_count": dv_count,
            })

        # Defined names (named ranges)
        defined_names: list[str] = []
        try:
            for dn in wb.defined_names:
                defined_names.append(dn if isinstance(dn, str) else str(dn))
        except Exception:
            pass

        return {
            "volume_filepath": filepath,
            "filename": os.path.basename(filepath),
            "file_size_bytes": file_stat.st_size,
            "file_last_modified": _iso(getattr(props, "modified", None)),
            "file_created": _iso(getattr(props, "created", None)),
            "author": getattr(props, "creator", None),
            "last_modified_by": getattr(props, "lastModifiedBy", None),
            "title": getattr(props, "title", None),
            "description": getattr(props, "description", None),
            "is_password_protected": bool(
                getattr(wb, "security", None)
                and getattr(wb.security, "lockStructure", False)
            ),
            "sheet_count": len(wb.sheetnames),
            "sheets": sheets_info,
            "defined_names": defined_names,
        }
    finally:
        wb.close()


@mcp.tool()
def read_data_from_excel(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: str | None = None,
    preview_only: bool = False,
    max_rows: int = 50,
) -> dict[str, Any]:
    """[Step 3 — 실제 값 샘플] 지정한 범위의 셀 값을 2D 리스트로 반환한다.

    profile_sheet 결과로 데이터 영역과 컬럼 구조를 파악한 뒤, LLM 이 실제 값을
    직접 눈으로 확인해야 할 때 호출. 범위 지정으로 토큰 절약.

    Args:
        filepath: UC Volume 절대 경로 (/Volumes/...). 로컬/workspace 경로 불가.
        sheet_name: 시트 이름.
        start_cell: 시작 셀 (기본 A1).
        end_cell: 종료 셀. 미지정 시 시트의 used range 끝까지.
        preview_only: True 면 상위 max_rows 행만 반환.
        max_rows: preview_only=True 일 때 최대 반환 행 수 (기본 50).
    """
    local = resolve(filepath)
    rows = read_excel_range(local, sheet_name, start_cell, end_cell, preview_only)
    truncated = False
    if preview_only and len(rows) > max_rows:
        rows = rows[:max_rows]
        truncated = True
    return {
        "filepath": filepath,
        "sheet": sheet_name,
        "start_cell": start_cell,
        "end_cell": end_cell,
        "preview_only": preview_only,
        "truncated": truncated,
        "row_count": len(rows),
        "rows": rows,
    }


@mcp.tool()
def get_merged_cells(filepath: str, sheet_name: str) -> dict[str, Any]:
    """지정한 시트의 병합 셀 범위 목록을 반환한다.

    haris-musa/excel-mcp-server 의 get_merged_cells 와 동일한 기능을 UC Volume 경로
    인터페이스로 노출한다. get_workbook_metadata 는 모든 시트의 병합셀을 한 번에
    돌려주지만, 이 tool 은 단일 시트 focused query 용도로 더 가볍다.

    Args:
        filepath: UC Volume 절대 경로 (/Volumes/...).
        sheet_name: 시트 이름.
    """
    local = resolve(filepath)
    ranges = _get_merged_ranges_impl(local, sheet_name)
    return {
        "filepath": filepath,
        "sheet": sheet_name,
        "merged_ranges": [str(r) for r in ranges],
        "merged_cell_count": len(ranges),
    }


@mcp.tool()
def get_data_validation_info(filepath: str, sheet_name: str) -> dict[str, Any]:
    """지정한 시트의 데이터 검증(Data Validation) 규칙을 모두 반환한다.

    haris-musa/excel-mcp-server 의 get_data_validation_info 와 동일한 기능.
    어떤 셀 범위에 어떤 종류(list/whole/decimal/date/textLength/custom)의 검증이
    걸려 있는지 LLM 이 한 번에 파악할 수 있도록 구조화된 결과를 돌려준다.

    Args:
        filepath: UC Volume 절대 경로 (/Volumes/...).
        sheet_name: 시트 이름.
    """
    local = resolve(filepath)
    wb = load_workbook(local, read_only=False, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            return {
                "filepath": filepath,
                "sheet": sheet_name,
                "error": f"Sheet '{sheet_name}' not found",
                "validations": [],
                "validation_count": 0,
            }
        ws = wb[sheet_name]
        validations = _get_all_validation_ranges_impl(ws) or []
        return {
            "filepath": filepath,
            "sheet": sheet_name,
            "validations": validations,
            "validation_count": len(validations),
        }
    finally:
        wb.close()


@mcp.tool()
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> dict[str, Any]:
    """Excel 수식 문법을 검증한다 (실제 시트에 적용하지 않음).

    haris-musa/excel-mcp-server 의 validate_formula_syntax 와 동일한 read-only 검증.
    수식이 문법적으로 올바른지, 셀 참조가 유효한지 등을 확인한다.

    Args:
        filepath: UC Volume 절대 경로 (/Volumes/...).
        sheet_name: 시트 이름.
        cell: 수식이 들어갈 셀 (예: "B2").
        formula: 검증할 수식 (예: "=SUM(A1:A10)").
    """
    local = resolve(filepath)
    try:
        result = _validate_formula_impl(local, sheet_name, cell, formula)
        return {
            "filepath": filepath,
            "sheet": sheet_name,
            "cell": cell,
            "formula": formula,
            "valid": True,
            "message": result.get("message", "OK"),
        }
    except Exception as e:
        return {
            "filepath": filepath,
            "sheet": sheet_name,
            "cell": cell,
            "formula": formula,
            "valid": False,
            "message": f"Error: {e}",
        }


@mcp.tool()
def validate_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str | None = None,
) -> dict[str, Any]:
    """지정한 범위가 시트 안에 존재하고 형식이 올바른지 검증한다.

    haris-musa/excel-mcp-server 의 validate_excel_range 와 동일한 read-only 검증.

    Args:
        filepath: UC Volume 절대 경로 (/Volumes/...).
        sheet_name: 시트 이름.
        start_cell: 시작 셀.
        end_cell: 종료 셀 (생략 시 단일 셀로 검증).
    """
    local = resolve(filepath)
    range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
    try:
        result = _validate_range_impl(local, sheet_name, range_str)
        return {
            "filepath": filepath,
            "sheet": sheet_name,
            "range": range_str,
            "valid": True,
            "message": result.get("message", "OK"),
        }
    except Exception as e:
        return {
            "filepath": filepath,
            "sheet": sheet_name,
            "range": range_str,
            "valid": False,
            "message": f"Error: {e}",
        }


# ---------------------------------------------------------------------------
# profile_sheet 를 위한 내부 유틸
# ---------------------------------------------------------------------------
_SPARK_TYPE_MAP = {
    "int": "BIGINT",
    "float": "DOUBLE",
    "bool": "BOOLEAN",
    "datetime": "TIMESTAMP",
    "date": "DATE",
    "time": "STRING",  # Spark 에 TIME 타입 없음
    "str": "STRING",
    "other": "STRING",
    "null": "STRING",
    "unknown": "STRING",
}


def _sanitize_column_name(name: Any) -> str:
    """Spark SQL 에서 쓸 수 있는 안전한 컬럼명으로 변환."""
    s = "" if name is None else str(name).strip()
    if not s:
        return ""
    # 영문·숫자·한글·언더스코어 외 모두 _ 로
    s = re.sub(r"[^\w가-힣]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        return ""
    if s[0].isdigit():
        s = "c_" + s
    return s[:128]


def _row_fill_counts(ws) -> list[int]:
    """각 row 의 non-null 셀 수.

    openpyxl 기본 동작상 병합 셀의 anchor(좌상단) 외 칸은 .value 가 None 으로 보이므로,
    원본 그대로 카운트하면 병합 영역이 sparse 로 잘못 분류된다. 이 함수는 병합 anchor
    값을 영역 전체에 풀어서 점유로 인정한다 — _detect_data_region 의 헤더/데이터 경계
    탐지가 한국형 Excel(상단 병합 제목, 다단 헤더, 부분 병합 합계 등)에서도 견고해진다.
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    merged_anchor: dict[tuple[int, int], Any] = {}
    for mr in ws.merged_cells.ranges:
        anchor_val = ws.cell(row=mr.min_row, column=mr.min_col).value
        if anchor_val in (None, ""):
            continue
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_anchor[(r, c)] = anchor_val

    counts = []
    for r in range(1, max_row + 1):
        count = 0
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val in (None, "") and (r, c) in merged_anchor:
                val = merged_anchor[(r, c)]
            if val not in (None, ""):
                count += 1
        counts.append(count)
    return counts


def _detect_data_region(ws) -> dict[str, Any]:
    """시트의 실제 데이터 영역 범위를 휴리스틱으로 탐지.

    기본 아이디어: '조밀한(non-null 많은)' 연속 블록이 데이터 영역.
    그 위의 적게 채워진 행들 = 제목/메타, 밑의 적게 채워진 행들 = 소계/합계.
    """
    row_fill = _row_fill_counts(ws)
    if not row_fill:
        return {
            "leading_title_rows": [],
            "header_row": None,
            "data_start_row": None,
            "data_end_row": None,
            "trailing_summary_rows": [],
            "row_fill_counts": [],
        }

    max_fill = max(row_fill)
    dense_thr = max(2, int(max_fill * 0.6))
    min_thr = max(1, int(max_fill * 0.3))

    # 조밀한 첫 행 = 헤더 후보
    header_row = None
    for i, f in enumerate(row_fill, start=1):
        if f >= dense_thr:
            header_row = i
            break

    if header_row is None:
        # 전체가 희소 → 첫 non-empty 행을 헤더로
        for i, f in enumerate(row_fill, start=1):
            if f > 0:
                header_row = i
                break

    data_start = (header_row + 1) if header_row else None

    # data_end: 헤더 이후로 조밀도가 min_thr 이상인 연속 구간 끝
    data_end = None
    if data_start and data_start <= len(row_fill):
        last_dense = None
        for i in range(data_start, len(row_fill) + 1):
            f = row_fill[i - 1]
            if f >= min_thr:
                last_dense = i
            elif last_dense is not None and f < min_thr / 2:
                # 연속 구간이 끊긴 지점에서 멈춤
                break
        data_end = last_dense

    leading_titles = list(range(1, header_row)) if header_row and header_row > 1 else []
    trailing = []
    if data_end and data_end < len(row_fill):
        for i in range(data_end + 1, len(row_fill) + 1):
            if row_fill[i - 1] > 0:
                trailing.append(i)

    return {
        "leading_title_rows": leading_titles,
        "header_row": header_row,
        "data_start_row": data_start,
        "data_end_row": data_end,
        "trailing_summary_rows": trailing,
        "row_fill_counts": row_fill,
    }


def _detect_header_rows(ws, header_row_candidate: int | None) -> list[int]:
    """다단 헤더 감지 — 헤더 후보 위쪽으로 3행 이내에 가로 병합이 있으면 상위 헤더로 간주."""
    if not header_row_candidate:
        return []
    rows = [header_row_candidate]
    for merged in ws.merged_cells.ranges:
        # 가로 병합 (한 행) 이면서 header_row_candidate 위쪽 3행 이내에 끝나는 것
        if merged.min_row == merged.max_row and merged.max_row < header_row_candidate \
                and merged.max_row >= max(1, header_row_candidate - 3):
            if merged.min_row not in rows:
                rows.append(merged.min_row)
    return sorted(rows)


def _flatten_header_row(ws, header_rows: list[int], max_col: int) -> list[str]:
    """여러 헤더 행을 '__' 로 연결해 하나의 컬럼명 리스트로 만든다."""
    if not header_rows:
        return [f"col_{c}" for c in range(1, max_col + 1)]

    # 병합 anchor 값 lookup
    merge_anchor: dict[tuple[int, int], Any] = {}
    for merged in ws.merged_cells.ranges:
        anchor_val = ws.cell(row=merged.min_row, column=merged.min_col).value
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                merge_anchor[(r, c)] = anchor_val

    headers: list[str] = []
    last_parts_per_row: dict[int, str] = {}
    for c in range(1, max_col + 1):
        parts: list[str] = []
        for hr in header_rows:
            val = ws.cell(row=hr, column=c).value
            if val in (None, "") and (hr, c) in merge_anchor:
                val = merge_anchor[(hr, c)]
            s = "" if val is None else str(val).strip()
            if s:
                last_parts_per_row[hr] = s
                parts.append(s)
            elif hr != header_rows[-1] and hr in last_parts_per_row:
                # 상위 헤더에서 빈 셀이면 직전 값 유지 (다단 헤더 가로 확장 방식)
                parts.append(last_parts_per_row[hr])
        name = "__".join(p for p in parts if p) or f"col_{c}"
        headers.append(name)
    return headers


def _profile_column(values: list[Any]) -> dict[str, Any]:
    """단일 컬럼 값 리스트를 받아 타입·통계·경고를 반환."""
    import datetime as _dt

    total = len(values)
    non_null = [v for v in values if v is not None and v != ""]
    null_count = total - len(non_null)

    type_counts: dict[str, int] = {}
    for v in non_null:
        if isinstance(v, bool):
            t = "bool"
        elif isinstance(v, int):
            t = "int"
        elif isinstance(v, float):
            t = "float"
        elif isinstance(v, _dt.datetime):
            t = "datetime"
        elif isinstance(v, _dt.date):
            t = "date"
        elif isinstance(v, _dt.time):
            t = "time"
        elif isinstance(v, str):
            t = "str"
        else:
            t = "other"
        type_counts[t] = type_counts.get(t, 0) + 1

    if not non_null:
        inferred = "null"
    else:
        # int + float 공존 → float 로 통합
        t_set = set(type_counts)
        if t_set <= {"int", "float"} and "float" in t_set:
            inferred = "float"
        elif len(t_set) == 1:
            inferred = next(iter(t_set))
        else:
            inferred = max(type_counts.items(), key=lambda kv: kv[1])[0]

    is_mixed = len([k for k in type_counts]) > 1 and not (
        set(type_counts) <= {"int", "float"}
    )

    # unique / samples
    try:
        uniq = set(non_null)
        unique_count = len(uniq)
    except TypeError:
        unique_count = None

    sample_values: list[str] = []
    seen: set = set()
    for v in non_null:
        try:
            if v in seen:
                continue
            seen.add(v)
        except TypeError:
            pass
        sample_values.append(str(v)[:100])
        if len(sample_values) >= 5:
            break

    # min / max
    mn: Any = None
    mx: Any = None
    try:
        if inferred in ("int", "float"):
            nums = [v for v in non_null if isinstance(v, (int, float)) and not isinstance(v, bool)]
            if nums:
                mn = min(nums)
                mx = max(nums)
        elif inferred in ("datetime", "date"):
            dts = [v for v in non_null if hasattr(v, "year")]
            if dts:
                mn = min(dts).isoformat()
                mx = max(dts).isoformat()
        elif inferred == "str":
            lens = [len(v) for v in non_null if isinstance(v, str)]
            if lens:
                mn = min(lens)  # 최소 길이
                mx = max(lens)  # 최대 길이
    except Exception:
        pass

    warnings: list[str] = []
    if is_mixed:
        warnings.append(f"타입 혼재: {type_counts}")
    if total > 0:
        null_pct = 100.0 * null_count / total
        if null_pct >= 50:
            warnings.append(f"null 비율 {null_pct:.1f}%")
    if inferred == "null":
        warnings.append("표본 전체가 비어 있음 — 실제 컬럼인지 재확인")

    return {
        "inferred_type": inferred,
        "spark_type": _SPARK_TYPE_MAP.get(inferred, "STRING"),
        "null_count": null_count,
        "null_pct": round(100.0 * null_count / total, 2) if total else 0.0,
        "unique_count": unique_count,
        "unique_pct": round(100.0 * unique_count / len(non_null), 2) if unique_count is not None and non_null else None,
        "min": mn if isinstance(mn, (int, float, str)) else str(mn) if mn is not None else None,
        "max": mx if isinstance(mx, (int, float, str)) else str(mx) if mx is not None else None,
        "type_distribution": type_counts,
        "is_mixed_type": is_mixed,
        "sample_values": sample_values,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Tools
# ---------------------------------------------------------------------------
def _profile_sheet_impl(
    filepath: str,
    sheet_name: str,
    max_sample_rows: int = 200,
) -> dict[str, Any]:
    """profile_sheet 의 핵심 로직. @mcp.tool() 래핑과 독립적으로 호출 가능하게 분리."""
    local = resolve(filepath)
    wb = load_workbook(local, read_only=False, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]
        max_col = ws.max_column or 0

        # 1. 데이터 영역 탐지
        region = _detect_data_region(ws)
        header_rows = _detect_header_rows(ws, region["header_row"])
        flat_headers = _flatten_header_row(ws, header_rows, max_col)
        sanitized = [_sanitize_column_name(h) or f"col_{i+1}" for i, h in enumerate(flat_headers)]

        data_start = region["data_start_row"]
        data_end = region["data_end_row"]

        # 2. 컬럼 프로파일
        columns_info: list[dict[str, Any]] = []
        if data_start and data_end and max_col > 0:
            sample_start = data_start
            sample_end = min(data_end, data_start + max_sample_rows - 1)
            for c in range(1, max_col + 1):
                values = [ws.cell(row=r, column=c).value for r in range(sample_start, sample_end + 1)]
                prof = _profile_column(values)
                columns_info.append({
                    "index": c - 1,
                    "letter": get_column_letter(c),
                    "original_header": flat_headers[c - 1] if c <= len(flat_headers) else None,
                    "sanitized_name": sanitized[c - 1] if c <= len(sanitized) else f"col_{c}",
                    **prof,
                })

        # 3. 수식 컬럼 탐지 (상위 20행 샘플)
        formula_columns: list[str] = []
        if data_start and max_col > 0:
            sample_rows_for_formula = min(20, (data_end - data_start + 1) if data_end else 20)
            for c in range(1, max_col + 1):
                has_formula = False
                for r in range(data_start, data_start + sample_rows_for_formula):
                    cell = ws.cell(row=r, column=c)
                    if getattr(cell, "data_type", None) == "f":
                        has_formula = True
                        break
                if has_formula:
                    formula_columns.append(get_column_letter(c))

        # 4. Excel feature 요약
        merged = [str(r) for r in ws.merged_cells.ranges]
        dv_list = getattr(ws, "data_validations", None)
        dv_rules = []
        if dv_list is not None:
            for dv in dv_list.dataValidation:
                dv_rules.append({
                    "type": getattr(dv, "type", None),
                    "formula1": getattr(dv, "formula1", None),
                    "ranges": [str(r) for r in dv.sqref.ranges] if getattr(dv, "sqref", None) else [],
                })
        hidden_rows = [r for r, dim in ws.row_dimensions.items() if getattr(dim, "hidden", False)]
        hidden_cols = [c for c, dim in ws.column_dimensions.items() if getattr(dim, "hidden", False)]

        excel_features = {
            "merged_cells": merged[:30],
            "merged_cells_truncated": len(merged) > 30,
            "merged_cell_count": len(merged),
            "has_formulas": bool(formula_columns),
            "formula_columns": formula_columns,
            "frozen_panes": ws.freeze_panes,
            "hidden_rows": hidden_rows,
            "hidden_columns": hidden_cols,
            "data_validation_rule_count": len(dv_rules),
            "data_validation_rules": dv_rules,
            "has_charts": bool(getattr(ws, "_charts", None)),
            "has_pivot_tables": bool(getattr(ws, "_pivots", None)),
        }

        # 5. Suggestions
        suggestions: list[str] = []
        if region["leading_title_rows"]:
            suggestions.append(
                f"상단 rows {region['leading_title_rows']} 은 제목/메타 행으로 보임 — 건너뛸 것"
            )
        if len(header_rows) > 1:
            suggestions.append(
                f"다단 헤더 감지 (rows {header_rows}) — '__' 구분자로 flatten 됨"
            )
        if region["trailing_summary_rows"]:
            suggestions.append(
                f"하단 rows {region['trailing_summary_rows']} 은 소계/합계로 보임 — 적재 시 제외 권장"
            )
        if formula_columns:
            suggestions.append(
                f"컬럼 {formula_columns} 에 수식 존재 — derived/계산식 컬럼인지 확인"
            )
        for col in columns_info:
            if col["warnings"]:
                suggestions.append(
                    f"컬럼 {col['sanitized_name']}({col['letter']}): {'; '.join(col['warnings'])}"
                )
        for col in columns_info:
            if col.get("unique_pct") == 100.0 and col["inferred_type"] in ("int", "str"):
                suggestions.append(
                    f"컬럼 {col['sanitized_name']} 은 unique 100% — Primary Key 후보"
                )
                break

        return {
            "volume_filepath": filepath,
            "sheet": sheet_name,
            "layout": {
                "leading_title_rows": region["leading_title_rows"],
                "header_rows": header_rows,
                "data_start_row": data_start,
                "data_end_row": data_end,
                "trailing_summary_rows": region["trailing_summary_rows"],
                "detected_total_data_rows": (data_end - data_start + 1) if (data_start and data_end) else 0,
                "max_row": ws.max_row,
                "max_column": max_col,
            },
            "headers": {
                "is_multi_row": len(header_rows) > 1,
                "header_row_count": len(header_rows),
                "original": flat_headers,
                "sanitized": sanitized,
            },
            "columns": columns_info,
            "excel_features": excel_features,
            "suggestions": suggestions,
        }
    finally:
        wb.close()


@mcp.tool()
def inspect_workbook(
    filepath: str,
    sheet_names: list[str] | None = None,
    max_sample_rows: int = 200,
) -> dict[str, Any]:
    """[초기 탐색 + 심층 분석 통합] Workbook 메타와 시트별 심층 프로파일을 한 번에 반환.

    Excel 파일을 처음 볼 때 호출. 워크북 전체 구조 (시트 목록, 메타, 보호 여부 등)와
    각 시트의 데이터 영역/헤더/컬럼 프로파일을 함께 받아 후속 전처리 계획 수립에 사용.

    반환 구조:
      - workbook: 파일 메타 (작성자, 크기, 시트 요약 등 — 기존 get_workbook_metadata 결과)
      - sheets: { sheet_name: 심층 프로파일 결과 (layout/headers/columns/excel_features/suggestions) }

    Args:
        filepath: UC Volume 절대 경로 (/Volumes/<catalog>/<schema>/<volume>/<file>.xlsx).
        sheet_names: 심층 프로파일링할 시트 이름 리스트. None 이면 모든 시트.
        max_sample_rows: 컬럼 통계에 사용할 최대 데이터 행 수 (기본 200).

    Note:
        - 큰 워크북에서 모든 시트를 한 번에 프로파일링하면 시간이 걸릴 수 있다.
          관심 시트만 지정하려면 sheet_names 인자 사용.
        - 후속 단계: read_data_from_excel 로 실제 값 샘플링,
          필요 시 suggest_delta_schema 로 DDL 생성.
    """
    workbook_meta = _get_workbook_metadata_impl(filepath)

    target_sheets = sheet_names if sheet_names else [s["name"] for s in workbook_meta.get("sheets", [])]

    sheets_profile: dict[str, Any] = {}
    for name in target_sheets:
        try:
            sheets_profile[name] = _profile_sheet_impl(filepath, name, max_sample_rows)
        except Exception as e:
            sheets_profile[name] = {"error": str(e)}

    return {
        "workbook": workbook_meta,
        "sheets": sheets_profile,
    }


@mcp.tool()
def suggest_delta_schema(
    filepath: str,
    sheet_name: str,
    table_name: str | None = None,
    catalog: str | None = None,
    schema: str | None = None,
) -> dict[str, Any]:
    """[Step 4 — 최종 DDL 생성] profile_sheet 결과를 바탕으로 Spark SQL CREATE TABLE 문과
    전처리 단계를 제안한다.

    내부적으로 profile_sheet 를 호출하므로 별도로 profile_sheet 를 먼저 부르지
    않아도 된다. 단, 분석 결과를 사용자가 한 번 검토한 뒤 호출하는 것을 권장.

    Args:
        filepath: UC Volume 절대 경로 (/Volumes/...).
        sheet_name: 시트 이름.
        table_name: 생성할 테이블 이름. 미지정 시 sheet name sanitize 로 자동 생성.
        catalog: 타겟 UC catalog (선택). 지정 시 full name 에 포함.
        schema: 타겟 UC schema (선택). 지정 시 full name 에 포함.

    Returns:
        suggested_full_name / ddl / column_mapping / preprocessing_steps /
        primary_key_candidates / partition_candidates / warnings
    """
    profile = _profile_sheet_impl(filepath, sheet_name)

    cols = profile["columns"]
    layout = profile["layout"]
    headers = profile["headers"]

    tname = table_name or _sanitize_column_name(sheet_name).lower() or "excel_sheet"
    parts = [p for p in [catalog, schema, tname] if p]
    full_name = ".".join(parts) if parts else tname

    # DDL — 중복 컬럼명 대비 dedup
    seen: dict[str, int] = {}
    deduped_cols: list[dict[str, Any]] = []
    for col in cols:
        base = col["sanitized_name"] or f"col_{col['index']+1}"
        n = seen.get(base, 0)
        final = base if n == 0 else f"{base}_{n}"
        seen[base] = n + 1
        new_col = dict(col)
        new_col["final_column_name"] = final
        deduped_cols.append(new_col)

    col_defs = [f"  `{c['final_column_name']}` {c['spark_type']}" for c in deduped_cols]
    ddl = (
        f"CREATE TABLE IF NOT EXISTS {full_name} (\n"
        + ",\n".join(col_defs)
        + "\n) USING DELTA"
    )

    column_mapping = [
        {
            "excel_letter": c["letter"],
            "original_header": c["original_header"],
            "delta_column": c["final_column_name"],
            "spark_type": c["spark_type"],
            "inferred_python_type": c["inferred_type"],
            "nullable": c["null_count"] > 0,
            "unique_pct": c.get("unique_pct"),
            "warnings": c.get("warnings", []),
        }
        for c in deduped_cols
    ]

    steps: list[str] = []
    if layout["leading_title_rows"]:
        steps.append(
            f"1. 상단 rows {layout['leading_title_rows']} 건너뛰기 (제목/메타)"
        )
    if headers["is_multi_row"]:
        steps.append(
            f"2. 다단 헤더 rows {layout['header_rows']} 를 '__' 구분자로 flatten"
        )
    if layout["data_start_row"] and layout["data_end_row"]:
        steps.append(
            f"3. 데이터 영역만 추출: rows {layout['data_start_row']} ~ {layout['data_end_row']}"
        )
    if layout["trailing_summary_rows"]:
        steps.append(
            f"4. 하단 rows {layout['trailing_summary_rows']} 제외 (소계/합계)"
        )
    for c in deduped_cols:
        if c.get("is_mixed_type"):
            steps.append(
                f"5. 컬럼 {c['final_column_name']} 은 타입 혼재 — 정규화 필요 (distribution: {c['type_distribution']})"
            )
    if not steps:
        steps.append("특별한 전처리 없이 바로 적재 가능해 보임")

    pk_candidates = [
        c["final_column_name"]
        for c in deduped_cols
        if c.get("unique_pct") == 100.0 and c["inferred_type"] in ("int", "str")
    ]
    partition_candidates = [
        c["final_column_name"]
        for c in deduped_cols
        if c["inferred_type"] in ("date", "datetime")
    ]

    all_warnings = [w for c in deduped_cols for w in c.get("warnings", [])]

    return {
        "volume_filepath": filepath,
        "sheet": sheet_name,
        "suggested_table_name": tname,
        "suggested_full_name": full_name,
        "ddl": ddl,
        "column_mapping": column_mapping,
        "preprocessing_steps": steps,
        "primary_key_candidates": pk_candidates,
        "partition_candidates": partition_candidates,
        "warnings": all_warnings,
    }


# ---------------------------------------------------------------------------
# CORS configuration
# ---------------------------------------------------------------------------
# Genie Code (Databricks workspace UI) 에서 이 MCP 서버를 호출하려면
# 브라우저가 먼저 `OPTIONS /mcp` preflight 요청을 보내는데,
# FastMCP 기본 Starlette 앱은 OPTIONS 에 대해 405 를 반환한다.
# CORSMiddleware 를 추가하여 preflight 를 204 로 처리하고
# 적절한 Access-Control-* 헤더를 리턴하도록 한다.
#
# 공식 가이드: https://docs.databricks.com/aws/en/genie-code/mcp
#   - stateless_http=True 로 세션리스 HTTP 모드 사용
#   - workspace 도메인을 allow_origins 에 포함
# 워크스페이스 호스트는 deploy.sh 가 DATABRICKS_HOST 환경변수로 주입한다.
# 비어 있으면 allow_origin_regex 가 *.cloud.databricks.com 등을 커버한다.
_workspace_host = os.environ.get("DATABRICKS_HOST", "").rstrip("/")
_DEFAULT_ALLOWED_ORIGINS = [_workspace_host] if _workspace_host else []

# 환경변수로 추가 origin 주입 가능 (콤마 구분, 선택)
_extra = os.environ.get("MCP_ALLOWED_ORIGINS", "").strip()
ALLOWED_ORIGINS = _DEFAULT_ALLOWED_ORIGINS + (
    [o.strip() for o in _extra.split(",") if o.strip()] if _extra else []
)

CORS_MIDDLEWARE = Middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    # 동일 계정 내 다른 databricks workspace 도메인도 허용
    allow_origin_regex=r"https://.*\.(cloud\.databricks\.com|azuredatabricks\.net|databricksapps\.com)",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["mcp-session-id"],
)


# ---------------------------------------------------------------------------
# Keep-alive ASGI middleware
# ---------------------------------------------------------------------------
# Databricks Apps reverse proxy 가 long-lived SSE connection 을 ~5 분 idle 후
# 종료해 클라이언트가 새로 고침해야만 다시 동작하는 증상을 회피한다.
# text/event-stream 응답에 한해 KEEPALIVE_INTERVAL_SECONDS 마다 SSE comment
# (`: keepalive\n\n`) 를 본문에 주입한다. 클라이언트는 comment line 을 무시하므로
# MCP 프로토콜에는 영향을 주지 않으며, 연결은 idle 로 판단되지 않는다.
class KeepAliveASGIMiddleware:
    KEEPALIVE_INTERVAL_SECONDS = 240  # 4 분 (Apps 5 분 idle timeout 보다 짧게)

    def __init__(self, app):
        self.app = app

    async def __call__(self, scope, receive, send):
        if scope.get("type") != "http":
            await self.app(scope, receive, send)
            return

        is_sse = False
        send_lock = asyncio.Lock()
        done = asyncio.Event()
        keepalive_task: asyncio.Task | None = None
        interval = self.KEEPALIVE_INTERVAL_SECONDS

        async def keepalive_loop():
            while not done.is_set():
                try:
                    await asyncio.wait_for(done.wait(), timeout=interval)
                except asyncio.TimeoutError:
                    if not is_sse or done.is_set():
                        continue
                    async with send_lock:
                        try:
                            await send({
                                "type": "http.response.body",
                                "body": b": keepalive\n\n",
                                "more_body": True,
                            })
                        except Exception:
                            done.set()
                            return

        async def wrapped_send(message):
            nonlocal is_sse, keepalive_task
            mtype = message.get("type")
            if mtype == "http.response.start":
                for name, value in message.get("headers", []) or []:
                    if name.lower() == b"content-type" and b"text/event-stream" in value.lower():
                        is_sse = True
                        break
                async with send_lock:
                    await send(message)
                if is_sse and keepalive_task is None:
                    keepalive_task = asyncio.create_task(keepalive_loop())
                return
            if mtype == "http.response.body":
                async with send_lock:
                    await send(message)
                if not message.get("more_body", False):
                    done.set()
                return
            async with send_lock:
                await send(message)

        try:
            await self.app(scope, receive, wrapped_send)
        finally:
            done.set()
            if keepalive_task is not None:
                keepalive_task.cancel()
                try:
                    await keepalive_task
                except (asyncio.CancelledError, Exception):
                    pass


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main() -> None:
    port = int(os.environ.get("DATABRICKS_APP_PORT", os.environ.get("PORT", "8000")))
    host = os.environ.get("FASTMCP_HOST", "0.0.0.0")
    logger.info("Starting FastMCP (streamable-http, stateless) on %s:%d", host, port)
    logger.info("Cache dir: %s", CACHE_DIR)
    logger.info("CORS allow_origins: %s", ALLOWED_ORIGINS)
    mcp.run(
        transport="streamable-http",
        host=host,
        port=port,
        path="/mcp",
        stateless_http=True,
        # CORS 가장 바깥 → UserToken → KeepAlive 안쪽 (SSE 응답 본문에 주기적 ping 주입)
        middleware=[
            CORS_MIDDLEWARE,
            Middleware(UserTokenMiddleware),
            Middleware(KeepAliveASGIMiddleware),
        ],
    )


if __name__ == "__main__":
    main()
